#!/usr/bin/env python
"""
CUA-MASTER 송장 자동 처리 시스템
D:/주문취합/주문_배송 데이터와 연동
"""

import os
import sys
import pandas as pd
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('InvoiceProcessor')

class InvoiceProcessor:
    def __init__(self):
        self.base_path = Path("D:/주문취합/주문_배송")
        self.output_path = Path("C:/Users/8899y/CUA-MASTER/data/invoices")
        self.output_path.mkdir(parents=True, exist_ok=True)
        
        # 업체별 패턴
        self.vendors = {
            "반찬단지": ["반찬단지", "값진한끼"],
            "초심푸드": ["초심푸드", "카페24(초심푸드)"],
            "취영루": ["취영루", "로젠택배"],
            "인생": ["인생", "[인생]"],
            "모비딕": ["모비딕"],
            "BS": ["BS", "어썸"],
            "피자코리아": ["피자코리아"],
            "최씨남매": ["최씨남매"],
            "토마토소스": ["토마토소스", "카페24(토마토소스)"]
        }
        
    def find_latest_folder(self) -> Optional[Path]:
        """최신 날짜 폴더 찾기"""
        folders = [f for f in self.base_path.iterdir() if f.is_dir() and f.name.isdigit()]
        if not folders:
            return None
        return max(folders, key=lambda x: x.name)
        
    def process_order_file(self, file_path: Path) -> pd.DataFrame:
        """주문 파일 처리"""
        try:
            if file_path.suffix == '.csv':
                # 다양한 인코딩 시도
                for encoding in ['utf-8', 'cp949', 'euc-kr']:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        logger.info(f"CSV 파일 읽기 성공: {file_path.name} (encoding: {encoding})")
                        return df
                    except:
                        continue
                        
            elif file_path.suffix in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path)
                logger.info(f"Excel 파일 읽기 성공: {file_path.name}")
                return df
                
        except Exception as e:
            logger.error(f"파일 읽기 실패 {file_path}: {e}")
            
        return pd.DataFrame()
        
    def extract_tracking_numbers(self, df: pd.DataFrame) -> List[str]:
        """송장번호 추출"""
        tracking_numbers = []
        
        # 가능한 송장번호 컬럼명
        possible_columns = [
            '송장번호', '운송장번호', 'tracking', 'tracking_number',
            '택배송장번호', '배송번호', '운송장', '송장'
        ]
        
        for col in df.columns:
            if any(keyword in col.lower() for keyword in ['송장', 'tracking', '운송장', '배송']):
                if col in df.columns:
                    numbers = df[col].dropna().astype(str).tolist()
                    tracking_numbers.extend(numbers)
                    logger.info(f"송장번호 {len(numbers)}개 추출: {col}")
                    
        return list(set(tracking_numbers))  # 중복 제거
        
    def match_vendor(self, filename: str) -> Optional[str]:
        """파일명으로 업체 매칭"""
        for vendor, patterns in self.vendors.items():
            if any(pattern in filename for pattern in patterns):
                return vendor
        return None
        
    def process_daily_invoices(self, date_str: str) -> Dict:
        """일별 송장 처리"""
        folder_path = self.base_path / date_str
        if not folder_path.exists():
            logger.error(f"폴더 없음: {folder_path}")
            return {}
            
        result = {
            "date": date_str,
            "vendors": {},
            "total_orders": 0,
            "total_tracking": 0,
            "files_processed": []
        }
        
        # 송장 폴더 우선 처리
        invoice_folder = folder_path / "송장"
        if invoice_folder.exists():
            files = list(invoice_folder.glob("*.*"))
        else:
            files = list(folder_path.glob("*.*"))
            
        for file_path in files:
            if file_path.suffix in ['.csv', '.xlsx', '.xls']:
                vendor = self.match_vendor(file_path.name)
                if vendor:
                    df = self.process_order_file(file_path)
                    if not df.empty:
                        tracking_numbers = self.extract_tracking_numbers(df)
                        
                        if vendor not in result["vendors"]:
                            result["vendors"][vendor] = {
                                "files": [],
                                "tracking_numbers": [],
                                "order_count": 0
                            }
                            
                        result["vendors"][vendor]["files"].append(file_path.name)
                        result["vendors"][vendor]["tracking_numbers"].extend(tracking_numbers)
                        result["vendors"][vendor]["order_count"] += len(df)
                        
                        result["files_processed"].append(file_path.name)
                        result["total_orders"] += len(df)
                        result["total_tracking"] += len(tracking_numbers)
                        
        return result
        
    def create_upload_csv(self, tracking_data: Dict) -> Path:
        """Cafe24 업로드용 CSV 생성"""
        upload_data = []
        
        for vendor_data in tracking_data["vendors"].values():
            for tracking_number in vendor_data["tracking_numbers"]:
                if tracking_number and tracking_number != 'nan':
                    upload_data.append({
                        "주문번호": "",  # Cafe24에서 매칭
                        "송장번호": tracking_number,
                        "배송업체": "로젠택배",  # 기본값
                        "처리일자": tracking_data["date"]
                    })
                    
        if upload_data:
            df = pd.DataFrame(upload_data)
            output_file = self.output_path / f"upload_{tracking_data['date']}.csv"
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
            logger.info(f"업로드 CSV 생성: {output_file}")
            return output_file
            
        return None
        
    def generate_report(self, tracking_data: Dict) -> Path:
        """처리 보고서 생성"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "송장처리현황"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 헤더 작성
        headers = ["업체명", "파일수", "주문건수", "송장수", "처리율"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # 데이터 작성
        row = 2
        for vendor, data in tracking_data["vendors"].items():
            ws.cell(row=row, column=1, value=vendor)
            ws.cell(row=row, column=2, value=len(data["files"]))
            ws.cell(row=row, column=3, value=data["order_count"])
            ws.cell(row=row, column=4, value=len(data["tracking_numbers"]))
            
            if data["order_count"] > 0:
                rate = len(data["tracking_numbers"]) / data["order_count"] * 100
                ws.cell(row=row, column=5, value=f"{rate:.1f}%")
            else:
                ws.cell(row=row, column=5, value="0%")
                
            row += 1
            
        # 합계
        ws.cell(row=row+1, column=1, value="합계").font = Font(bold=True)
        ws.cell(row=row+1, column=2, value=len(tracking_data["files_processed"]))
        ws.cell(row=row+1, column=3, value=tracking_data["total_orders"])
        ws.cell(row=row+1, column=4, value=tracking_data["total_tracking"])
        
        if tracking_data["total_orders"] > 0:
            total_rate = tracking_data["total_tracking"] / tracking_data["total_orders"] * 100
            ws.cell(row=row+1, column=5, value=f"{total_rate:.1f}%")
            
        # 열 너비 조정
        for col in range(1, 6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
        # 저장
        report_file = self.output_path / f"report_{tracking_data['date']}.xlsx"
        wb.save(report_file)
        logger.info(f"보고서 생성: {report_file}")
        
        return report_file
        
    def process_today(self) -> Dict:
        """오늘 날짜 송장 처리"""
        today = datetime.now().strftime("%Y%m%d")
        return self.process_daily_invoices(today)
        
    def run(self, date_str: Optional[str] = None) -> Dict:
        """송장 처리 실행"""
        if not date_str:
            date_str = datetime.now().strftime("%Y%m%d")
            
        logger.info(f"송장 처리 시작: {date_str}")
        
        # 데이터 처리
        tracking_data = self.process_daily_invoices(date_str)
        
        if not tracking_data or not tracking_data["vendors"]:
            logger.warning(f"처리할 송장 데이터 없음: {date_str}")
            return {"status": "no_data", "date": date_str}
            
        # CSV 생성
        csv_file = self.create_upload_csv(tracking_data)
        
        # 보고서 생성  
        report_file = self.generate_report(tracking_data)
        
        # 결과 저장
        result = {
            "status": "success",
            "date": date_str,
            "summary": {
                "vendors": len(tracking_data["vendors"]),
                "total_orders": tracking_data["total_orders"],
                "total_tracking": tracking_data["total_tracking"],
                "processing_rate": (tracking_data["total_tracking"] / tracking_data["total_orders"] * 100) 
                                   if tracking_data["total_orders"] > 0 else 0
            },
            "files": {
                "csv": str(csv_file) if csv_file else None,
                "report": str(report_file)
            },
            "details": tracking_data
        }
        
        # JSON 저장
        json_file = self.output_path / f"result_{date_str}.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
            
        logger.info(f"처리 완료: {tracking_data['total_tracking']}개 송장")
        
        return result

if __name__ == "__main__":
    processor = InvoiceProcessor()
    
    # 명령줄 인자로 날짜 받기
    if len(sys.argv) > 1:
        date_str = sys.argv[1]
    else:
        date_str = None
        
    result = processor.run(date_str)
    print(json.dumps(result, ensure_ascii=False, indent=2))